#!/usr/bin/env python3
"""
Générateur d'XLSForm/CSV Kobo pour la soumission de prix agricoles.

Fonctionnalités:
- Extrait les listes (catégories, produits, localités, unités, langues) depuis l'API.
- Construit les feuilles survey/choices/settings pour refléter le formulaire "Soumettre un prix" du frontend.
- Applique les contraintes de validation (prix >= 0, date <= today(), commentaire <= 500, regex téléphone 01XXXXXXXX).
- Ajoute le filtrage des produits par catégorie via choice_filter et relevant.

Prérequis:
- Python 3.8+
- Packages: openpyxl (uniquement si sortie XLSX), requests
  Installation: pip install openpyxl requests

Utilisation:
- Par défaut (API locale, XLSX):
    python scripts/generate_kobo_xlsform.py --api-url http://localhost:5000/api

- Sortie CSV au lieu de XLSX:
    python scripts/generate_kobo_xlsform.py --csv --csv-dir scripts/output/kobo_price_submission_csv --api-url http://localhost:5000/api

- Spécifier le fichier XLSX de sortie:
    python scripts/generate_kobo_xlsform.py --output scripts/output/kobo_price_submission.xlsx --api-url http://localhost:5000/api

Notes:
- Le script lit les données via l'API backend (PostgreSQL/Supabase).
- Si vous souhaitez personnaliser les libellés, ajustez les fonctions build_* ci-dessous.
"""

import argparse
import os
import sys
import csv
from typing import List, Dict, Any, Optional


# SQLite supprimé: extraction uniquement via API


def fetch_from_api(api_url: str) -> Dict[str, List[Dict[str, Any]]]:
    try:
        import requests
    except ImportError:
        print("Le package 'requests' est requis pour l'extraction via API. Installez-le via: pip install requests", file=sys.stderr)
        raise

    def g(endpoint: str) -> List[Dict[str, Any]]:
        url = api_url.rstrip('/') + '/' + endpoint.lstrip('/')
        resp = requests.get(url, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        # Les routes renvoient { data: [...] } typiquement
        items = data.get('data') if isinstance(data, dict) else None
        return items or []

    return {
        "categories": g('product-categories'),
        "products": g('products'),
        "localities": g('localities'),
        "units": g('units'),
        "languages": g('languages'),
    }


def build_choices(data: Dict[str, List[Dict[str, Any]]]) -> List[List[str]]:
    """Construit les lignes de la feuille choices.
    Colonnes: list_name, name, label, [category_id pour products]
    """
    rows: List[List[str]] = [["list_name", "name", "label", "category_id", "symbol"]]

    # categories
    for c in data.get("categories", []):
        label = f"{c.get('name','')} ({c.get('type','')})".strip()
        rows.append(["categories", str(c.get("id")), label, "", ""])
    # option 'Autre'
    rows.append(["categories", "other", "Autre (spécifier)", "", ""])

    # products (avec colonne category_id pour le filtrage)
    for p in data.get("products", []):
        rows.append(["products", str(p.get("id")), str(p.get("name", "")), str(p.get("category_id")), ""])
    # option 'Autre' par catégorie existante (label inclut le nom de la catégorie)
    for c in data.get("categories", []):
        cid = str(c.get("id"))
        cname = str(c.get("name", "")).strip()
        if cid:
            other_label = f"Autre (spécifier) — {cname}" if cname else "Autre (spécifier)"
            rows.append(["products", f"other_{cid}", other_label, cid, ""])
    # option 'Autre' pour la catégorie 'other' (label explicite)
    rows.append(["products", "other", "Autre (spécifier) — Autre", "other", ""])

    # localities (pas d'option 'Autre' demandée)
    for l in data.get("localities", []):
        rows.append(["localities", str(l.get("id")), str(l.get("name", "")), "", ""])

    # units
    for u in data.get("units", []):
        label = str(u.get("name", ""))
        symbol = (u.get("symbol") or "").strip()
        if symbol:
            label = f"{label} ({symbol})"
        rows.append(["units", str(u.get("id")), label, "", symbol])
    rows.append(["units", "other", "Autre (spécifier)", "", ""])  # option 'Autre'

    # languages
    for lg in data.get("languages", []):
        rows.append(["languages", str(lg.get("id")), str(lg.get("name", "")), "", ""])
    rows.append(["languages", "other", "Autre (spécifier)", "", ""])  # option 'Autre'

    # category_types (statique: valeurs autorisées pour type de catégorie)
    rows.append(["category_types", "brut", "brut", "", ""])  # brut
    rows.append(["category_types", "transforme", "transformé", "", ""])  # transformé

    # source_type (statique)
    for st in ["producteur", "transformateur", "cooperative", "grossiste", "commercant", "autre"]:
        rows.append(["source_type", st, st, "", ""])

    # source_relation (statique)
    for sr in ["moi", "proche", "autre"]:
        rows.append(["source_relation", sr, sr, "", ""])

    return rows


def build_survey() -> List[List[str]]:
    """Construit les lignes de la feuille survey.
    Colonnes: type, name, label, required, constraint, constraint_message, relevant, appearance, hint, choice_filter, calculation
    """
    rows: List[List[str]] = [[
        "type", "name", "label", "required", "constraint", "constraint_message", "relevant", "appearance", "hint", "choice_filter", "calculation"
    ]]

    # Note introductive
    rows.append([
        "note",
        "intro_note",
        "Formulaire réservé aux contributeurs/autorisés. Note: au début du formulaire Kobo, renseignez votre nom d'utilisateur Lokali (ex: amitystokes-63hws7) dans le champ \"Nom d'utilisateur\". Cela permet de lier vos soumissions à votre compte. Pour plus de fiabilité, renseignez également votre identifiant Supabase (ID).",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        ""
    ])

    # Identifiant contributeur
    rows.append([
        "text", "submitted_username", "Votre nom d'utilisateur Lokali", "yes",
        "regex(., '^[A-Za-z0-9._-]{3,32}$')",
        "Utilisez 3–32 caractères: lettres, chiffres, . _ -",
        "", "", "Saisissez votre nom d'utilisateur Lokali (ex: amitystokes-63hws7).", "", ""
    ])
    # ID utilisateur optionnel (si username non-unique)
    rows.append([
        "text", "submitted_user_id", "Votre identifiant Supabase (ID)", "",
        "string-length(.) = 0 or string-length(.) >= 8",
        "Facultatif, mais recommandé pour améliorer la fiabilité de l'association compte-soumissions.",
        "", "", "Renseignez votre identifiant Supabase (ID) pour plus de fiabilité.", "", ""
    ])
    # Note configuration Kobo (sans duplication)
    rows.append([
        "note", "kobo_setup_note",
        "Configurez KoboCollect/KoboToolbox avec votre compte. Pour plus de fiabilité, renseignez également votre identifiant Supabase (ID).",
        "", "", "", "", "", "", "", ""
    ])

    # Catégorie
    rows.append(["select_one categories", "category_id", "Catégorie de produit", "yes", "", "", "", "search", "Recherchez par nom de catégorie.", "", ""])
    # Nouvelle catégorie si 'Autre'
    rows.append([
        "text", "new_category_name", "Nouvelle catégorie", "yes",
        "string-length(.) >= 2", "2 caractères minimum",
        "${category_id} = 'other'", "", "Écrivez le nom exact.", "", ""
    ])
    # Type de catégorie si 'Autre'
    rows.append([
        "select_one category_types", "new_category_type", "Type de la catégorie", "yes",
        "", "",
        "${category_id} = 'other'", "", "Sélectionnez: brut ou transformé.", "", ""
    ])

    # Produit (filtré par catégorie)
    rows.append([
        "select_one products", "product_id", "Produit", "yes",
        "", "", "${category_id} != ''", "search", "Filtré par catégorie; recherchez le produit.", "category_id=${category_id}", ""
    ])
    # Nouveau produit si 'Autre' (par catégorie) ou catégorie 'Autre'
    rows.append([
        "text", "new_product_name", "Nouveau produit", "yes",
        "string-length(.) >= 2", "2 caractères minimum",
        "${product_id} = 'other' or starts-with(${product_id}, 'other_') or ${category_id} = 'other'", "", "Écrivez le nom exact.", "", ""
    ])

    # Localité
    rows.append(["select_one localities", "locality_id", "Commune / Localité", "yes", "", "", "", "search", "Recherchez la commune/localité.", "", ""])

    # Sous-localité
    rows.append(["text", "sub_locality", "Village / Quartier / Hameau (optionnel)", "", "", "", "", "", "Décrivez la sous-localité précise si utile.", "", ""])

    # Prix
    rows.append([
        "decimal", "price", "Prix (FCFA)", "yes", ". >= 0", "Le prix doit être positif", "", "", "", "", ""
    ])

    # Unité
    rows.append(["select_one units", "unit_id", "Unité", "yes", "", "", "", "search", "Recherchez l’unité (ex: kg, l, unité).", "", ""])
    rows.append([
        "text", "new_unit_name", "Nouvelle unité", "yes",
        "string-length(.) >= 1", "Indiquez l'unité (ex: kg, l, unité)",
        "${unit_id} = 'other'", "", "Nom d'unité exact.", "", ""
    ])
    # Symbole de l'unité si 'Autre' (optionnel)
    rows.append([
        "text", "new_unit_symbol", "Symbole de l'unité (optionnel)", "",
        "", "",
        "${unit_id} = 'other'", "", "Ex: kg, l, t, u.", "", ""
    ])

    # Date
    rows.append([
        "date", "date", "Date", "yes", ". <= today()", "La date ne peut pas être future", "", "", "", "", ""
    ])

    # Commentaire
    rows.append([
        "text", "comment", "Commentaire (optionnel)", "", "string-length(.) <= 500", "500 caractères max", "", "", "Informations supplémentaires sur ce prix...", "", ""
    ])

    # Note sur la fiabilité des données via les informations de source
    rows.append([
        "note", "source_info_note",
        "Les informations de source permettent de vérifier la fiabilité des données.",
        "", "", "", "", "", "", "", ""
    ])

    # Source + type
    rows.append(["text", "source", "Nom de la source", "yes", "", "La source est requise pour assurer la fiabilité", "", "", "Ex: K. Diarra, marché X. Ces informations aident à vérifier la fiabilité des données.", "", ""])
    rows.append(["select_one source_type", "source_type", "Type de source", "yes", "", "Sélectionnez le type de source", "", "", "Ces informations aident à vérifier la fiabilité des données.", "", ""])
    # Si 'autre' est choisi, préciser (obligatoire quand pertinent)
    rows.append([
        "text", "source_type_other", "Préciser autre type de source", "yes",
        "string-length(.) >= 2", "2 caractères minimum",
        "selected(${source_type}, 'autre')", "", "Décrivez le type de source (ex: collecteur local)", "", ""
    ])

    # Contact
    rows.append(["text", "source_contact_name", "Nom du contact (optionnel)", "", "", "", "", "", "", "", ""])
    # Contrainte facultative: vide OU regex 01XXXXXXXX
    rows.append([
        "text", "source_contact_phone", "Téléphone du contact (01XXXXXXXX)", "", "(${source_contact_phone} = '' or regex(${source_contact_phone}, '^01\\d{8}$'))", "Format attendu: 01XXXXXXXX", "", "", "", "", ""
    ])
    rows.append(["select_one source_relation", "source_contact_relation", "Lien du contact (optionnel)", "", "", "", "", "", "Utile pour vérifier la fiabilité des données.", "", ""])

    # Géolocalisation
    rows.append(["geopoint", "gps", "Géolocalisation (idéalement ≤ 10 m)", "", "", "", "", "maps", "Activez la haute précision GPS et restez à ciel ouvert.", "", ""])

    # Langue de communication
    rows.append(["select_one languages", "source_language_id", "Langue de communication avec la source", "", "", "", "", "search", "Recherchez la langue de communication.", "", ""])
    rows.append([
        "text", "new_language_name", "Nouvelle langue", "yes",
        "string-length(.) >= 2", "2 caractères minimum",
        "${source_language_id} = 'other'", "", "Nom de langue exact.", "", ""
    ])

    # Flags de synchronisation API (calculés, non affichés)
    rows.append(["calculate", "sync_create_category", "", "", "", "", "", "", "", "", "if((${category_id} = 'other') and string-length(${new_category_name}) > 0 and string-length(${new_category_type}) > 0, 1, 0)"])
    rows.append(["calculate", "sync_create_product", "", "", "", "", "", "", "", "", "if((( ${product_id} = 'other') or starts-with(${product_id}, 'other_') or (${category_id} = 'other')) and string-length(${new_product_name}) > 0, 1, 0)"])
    rows.append(["calculate", "sync_create_unit", "", "", "", "", "", "", "", "", "if((${unit_id} = 'other') and string-length(${new_unit_name}) > 0, 1, 0)"])
    rows.append(["calculate", "sync_create_language", "", "", "", "", "", "", "", "", "if((${source_language_id} = 'other') and string-length(${new_language_name}) > 0, 1, 0)"])
    rows.append(["calculate", "submission_source", "", "", "", "", "", "", "", "", "'kobo'"])

    return rows


def build_settings() -> List[List[str]]:
    return [["form_title", "form_id", "version"], ["Soumettre un prix de produit agricole", "price_submission", "v1"]]


def write_workbook(output_path: str, survey: List[List[str]], choices: List[List[str]], settings: Optional[List[List[str]]] = None):
    try:
        from openpyxl import Workbook
    except ImportError:
        print("Le package 'openpyxl' est requis pour la sortie XLSX. Installez-le via: pip install openpyxl", file=sys.stderr)
        raise
    wb = Workbook()
    # Feuille survey
    ws_survey = wb.active
    ws_survey.title = "survey"
    for row in survey:
        ws_survey.append(row)

    # Feuille choices
    ws_choices = wb.create_sheet("choices")
    for row in choices:
        ws_choices.append(row)

    # Feuille settings
    if settings:
        ws_settings = wb.create_sheet("settings")
        for row in settings:
            ws_settings.append(row)

    # Création du répertoire si nécessaire
    out_dir = os.path.dirname(output_path)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    wb.save(output_path)
    print(f"XLSForm généré: {output_path}")


def write_csvs(output_dir: str, survey: List[List[str]], choices: List[List[str]], settings: Optional[List[List[str]]] = None):
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    def write_csv(path: str, rows: List[List[str]]):
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in rows:
                writer.writerow(row)

    survey_path = os.path.join(output_dir, 'survey.csv')
    choices_path = os.path.join(output_dir, 'choices.csv')
    write_csv(survey_path, survey)
    write_csv(choices_path, choices)
    if settings:
        settings_path = os.path.join(output_dir, 'settings.csv')
        write_csv(settings_path, settings)
    print(f"CSV générés: {survey_path}, {choices_path}, {(os.path.join(output_dir, 'settings.csv') if settings else 'aucun settings')} ")


def main():
    parser = argparse.ArgumentParser(description="Génère un XLSForm/CSV Kobo via API pour la soumission de prix")
    parser.add_argument("--api-url", default="http://localhost:5000/api", help="URL de base de l'API")
    parser.add_argument("--output", default=os.path.join("scripts", "output", "kobo_price_submission.xlsx"), help="Chemin de sortie du fichier XLS")
    parser.add_argument("--csv", action="store_true", help="Sortir au format CSV (survey/choices/settings)")
    parser.add_argument("--csv-dir", default=os.path.join("scripts", "output", "kobo_price_submission_csv"), help="Répertoire de sortie CSV")
    args = parser.parse_args()

    # Extraire les données (API uniquement)
    data = fetch_from_api(args.api_url)

    # Construire feuilles
    survey = build_survey()
    choices = build_choices(data)
    settings = build_settings()

    # Écrire XLSX ou CSV
    if args.csv:
        write_csvs(args.csv_dir, survey, choices, settings)
    else:
        write_workbook(args.output, survey, choices, settings)


if __name__ == "__main__":
    main()